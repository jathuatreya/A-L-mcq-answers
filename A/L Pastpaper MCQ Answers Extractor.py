import openpyxl

def extract_answer(file_path, year, question_number):
    # Load the Excel workbook
    workbook = openpyxl.load_workbook(file_path)
    
    # Assuming you have a sheet named 'Sheet1', change it accordingly
    sheet = workbook['Sheet1']
    
    # Find the column index based on the year
    year_column = None
    for col in sheet.iter_cols(min_col=2, max_col=sheet.max_column, min_row=1, max_row=1):
        for cell in col:
            if cell.value == year:
                year_column = col[0].column

    if year_column is None:
        print(f"No data found for year {year}")
        return
    
    # Find the row index based on the question number
    question_row = None
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=1):
        for cell in row:
            if cell.value == question_number:
                question_row = row[0].row

    if question_row is None:
        print(f"No data found for question number {question_number}")
        return
    
    
    answer = sheet.cell(row=question_row, column=year_column).value
    
    print(f"Answer for question {question_number} in year {year}: {answer}")

file_path = r'D:\language pratics\python project\chemistry mcq asn\che.xlsx'
print("welcome")
print("A/L pastpaper MCQ answers")
print("enter the year and the question number")
print("welcome")
user_year = int(input("Enter the year: "))
user_question_number = int(input("Enter the question number: "))

extract_answer(file_path, user_year, user_question_number)
