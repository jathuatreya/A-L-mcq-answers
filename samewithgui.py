import openpyxl
import tkinter as tk
from tkinter import ttk

DEFAULT_FILE_PATH = r'D:\language pratics\python project\chemistry mcq asn\che.xlsx'

def extract_answer(file_path, year, question_number):
    try:
        
        workbook = openpyxl.load_workbook(file_path)
        sheet = workbook['Sheet1']
        year_column = None
        for col in sheet.iter_cols(min_col=2, max_col=sheet.max_column, min_row=1, max_row=1):
            for cell in col:
                if cell.value == year:
                    year_column = col[0].column

        if year_column is None:
            return f"No data found for year {year}"
        
        # Find the row index based on the question number
        question_row = None
        for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=1):
            for cell in row:
                if cell.value == question_number:
                    question_row = row[0].row

        if question_row is None:
            return f"No data found for question number {question_number}"
        
        # Get the answer from the specified cell
        answer = sheet.cell(row=question_row, column=year_column).value
        
        return f"Answer for question {question_number} in year {year}: {answer}"
    
    except Exception as e:
        return f"Error: {str(e)}"

def on_entry_change(*args):
    try:
        user_year = int(year_entry.get())
        user_question_number = int(question_entry.get())
        
        result = extract_answer(DEFAULT_FILE_PATH, user_year, user_question_number)
        
        result_label.config(text=result)
    except ValueError:
        result_label.config(text="Invalid input. Please enter valid numbers.")

# Create the main window
root = tk.Tk()
root.title("chemistry mcq answer data extracter for 1979 to 2016")

# Create and place GUI elements
year_label = ttk.Label(root, text="Enter the year:")
year_label.grid(row=0, column=0, padx=5, pady=5)

year_entry = ttk.Entry(root)
year_entry.grid(row=0, column=1, padx=5, pady=5)
year_entry.bind('<KeyRelease>', on_entry_change)

question_label = ttk.Label(root, text="Enter the question number:")
question_label.grid(row=1, column=0, padx=5, pady=5)

question_entry = ttk.Entry(root)
question_entry.grid(row=1, column=1, padx=5, pady=5)
question_entry.bind('<KeyRelease>', on_entry_change)

result_label = ttk.Label(root, text="")
result_label.grid(row=2, column=0, columnspan=2, pady=10)

# Start the GUI event loop
root.mainloop()
